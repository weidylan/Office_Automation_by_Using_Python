#1.提取信息
from openpyxl import load_workbook
wb = load_workbook("data\领料总表.xlsx")
ws= wb.active
data = {} #用于储存提取的信息
for row in range(2, ws.max_row+1): #从第2行开始（第1行是标题）遍历工作表每一行，将数据提取出来
    work_order= ws['A' + str(row)].value #为工单号
    unit = ws['B' + str(row)].value #为物料单位
    material_pn = ws['C' + str(row)].value #为物料编号    
    qty = ws['F' + str(row)].value #批数量

    data.setdefault(work_order,{}) 
    data[work_order].setdefault(material_pn,{'单位':unit,
                                            '总数':0}) #先设定“总数”的初始值为0
    data[work_order][material_pn]['总数']+=int(qty) #让数量累加，确保所有批次的数量加总

#2.将提取的信息写入新建的Excel表
wb1 = load_workbook("data\领料汇总-模板.xlsx")
ws1= wb1.active

i=2 #计数器，从2开始（因为是从“领料汇总-模板”表中的第二行开始写）
for work_order in data.keys():
    for material in data[work_order].keys():
        ws1.cell(row=i,column=1).value=work_order #第1列为工单号
        ws1.cell(row=i,column=2).value=data[work_order][material]["单位"] #第2列为单位
        ws1.cell(row=i,column=3).value=material #第3列为物料编号
        ws1.cell(row=i,column=4).value=data[work_order][material]["总数"] #第4列为总数
        i+=1
wb1.save("data\领料汇总.xlsx")
